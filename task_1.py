import os
import re
import openpyxl
import requests

excel_path = "/Users/devbekasyl/Desktop/job_taks/Перечень.xlsx"

output_dir = "задание_1"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

hyperlink_pattern = r'HYPERLINK\("([^"]+)",'

wb = openpyxl.load_workbook(excel_path, data_only=False)  # data_only=False позволяет видеть формулы
sheet = wb.active

# Шаг 2. Обработка строк
for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Начиная со второй строки
    barcode = sheet.cell(row=row_idx, column=6).value  # Столбец F (штрих-код)

    if not barcode:  # Пропускаем строки без штрих-кода
        continue

    # Создаём папку для текущего штрих-кода
    barcode_folder = os.path.join(output_dir, str(barcode))
    os.makedirs(barcode_folder, exist_ok=True)

    # Обрабатываем столбцы с P по BB
    for col_idx in range(16, 55):  # Столбцы P-BB (16-й по 54-й)
        cell = sheet.cell(row=row_idx, column=col_idx)

        if cell.value and isinstance(cell.value, str) and cell.value.startswith("=HYPERLINK"):
            match = re.search(hyperlink_pattern, cell.value)
            if match:
                url = match.group(1)  # Извлекаем ссылку из формулы

                try:
                    # Загружаем файл
                    response = requests.get(url, verify=False)
                    if response.status_code == 200:
                        # Формируем имя файла
                        file_name = f"{barcode}-{col_idx - 15}.jpg"  # Нумерация начинается с 1
                        file_path = os.path.join(barcode_folder, file_name)

                        # Сохраняем файл
                        with open(file_path, 'wb') as f:
                            for chunk in response.iter_content(1024):
                                f.write(chunk)
                except Exception as e:
                    print(f"Ошибка при загрузке {url}: {e}")

print("Задача выполнена. Файлы сохранены в папке 'задание_1'.")